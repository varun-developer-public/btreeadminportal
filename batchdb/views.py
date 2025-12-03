from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.db import transaction
from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.views.decorators.http import require_POST, require_GET
from django.utils import timezone
from datetime import datetime, timedelta
from django.views.generic import ListView
from django.contrib.auth.mixins import LoginRequiredMixin

# REST Framework imports
from rest_framework import viewsets, status, filters
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from core.permissions import IsBatchCoordinator, IsStaff, IsTrainer
from django.views.decorators.csrf import csrf_exempt

# Model imports
from .models import (
    Batch, Course, Trainer, Student, BatchStudent,
    TransferRequest, BatchTransaction, TrainerHandover
)

# Serializer imports
from .serializers import (
    BatchSerializer, BatchDetailSerializer, BatchStudentSerializer,
    TransferRequestSerializer, TransferRequestDetailSerializer,
    TransferRequestApprovalSerializer, TransferRequestRejectionSerializer,
    TrainerHandoverSerializer, TrainerHandoverApprovalSerializer,
    TrainerHandoverRejectionSerializer, BatchTransactionSerializer,
    BatchTransactionDetailSerializer, StudentBatchHistorySerializer,
    StudentSerializer, TrainerSerializer
)

# Form imports
from .forms import BatchCreationForm, BatchUpdateForm, BatchFilterForm

import json
import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill

from coursedb.models import Course, CourseCategory
from trainersdb.models import Trainer
from studentsdb.models import Student
from placementdb.models import Placement

# Request Management API Endpoints

def batch_list(request):
    user = request.user
    if hasattr(user, 'trainer_profile'):
        batch_list = Batch.objects.filter(trainer=user.trainer_profile.trainer).order_by('-id')
    else:
        batch_list = Batch.objects.all().order_by('-id')
    form = BatchFilterForm(request.GET)

    if form.is_valid():
        query = form.cleaned_data.get('q')
        courses = form.cleaned_data.get('course')
        trainers = form.cleaned_data.get('trainer')
        statuses = form.cleaned_data.get('batch_status')
        start_date = form.cleaned_data.get('start_date')
        end_date = form.cleaned_data.get('end_date')
        time_slot = form.cleaned_data.get('time_slot')
        trainer_type = form.cleaned_data.get('trainer_type')
        percentage_min = form.cleaned_data.get('percentage_min')
        percentage_max = form.cleaned_data.get('percentage_max')

        if query:
            batch_list = batch_list.filter(
                Q(students__first_name__icontains=query) |
                Q(students__last_name__icontains=query) |
                Q(batch_id__icontains=query) |
                Q(course__course_name__icontains=query) |
                Q(trainer__name__icontains=query)
            ).distinct()
        
        if courses:
            batch_list = batch_list.filter(course__in=courses).distinct()
        
        if trainers:
            batch_list = batch_list.filter(trainer__in=trainers).distinct()

        if statuses:
            batch_list = batch_list.filter(batch_status__in=statuses).distinct()

        if start_date:
            batch_list = batch_list.filter(start_date__gte=start_date)

        if end_date:
            batch_list = batch_list.filter(end_date__lte=end_date)

        if time_slot:
            start_time_str, end_time_str = time_slot.split('-')
            start_time = datetime.strptime(start_time_str, '%H:%M').time()
            end_time = datetime.strptime(end_time_str, '%H:%M').time()
            batch_list = batch_list.filter(start_time=start_time, end_time=end_time)

        if trainer_type:
            batch_list = batch_list.filter(trainer__employment_type=trainer_type)

        if percentage_min is not None:
            batch_list = batch_list.filter(batch_percentage__gte=percentage_min)

        if percentage_max is not None:
            batch_list = batch_list.filter(batch_percentage__lte=percentage_max)

    paginator = Paginator(batch_list, 10)
    page = request.GET.get('page')

    try:
        batches = paginator.page(page)
    except PageNotAnInteger:
        batches = paginator.page(1)
    except EmptyPage:
        batches = paginator.page(paginator.num_pages)

    return render(request, 'batchdb/batch_list.html', {
        'batches': batches,
        'form': form
    })

@login_required
def create_batch(request):
    if request.method == 'POST':
        form = BatchCreationForm(request.POST)
        if form.is_valid():
            batch = form.save(commit=False)
            
            time_slot = form.cleaned_data.get('time_slot')
            if time_slot:
                batch.start_time, batch.end_time = time_slot

            days = form.cleaned_data.get('days')
            
            if not days:
                if batch.batch_type == 'WD':
                    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
                elif batch.batch_type == 'WE':
                    days = ['Saturday', 'Sunday']
                elif batch.batch_type == 'WDWE':
                    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
            
            batch.days = days
            batch.batch_status = 'YTS'  # Set default status
            batch.created_by = request.user
            batch.save()
            form.save_m2m()
            messages.success(request, "Batch created successfully.")
            return redirect('batchdb:batch_list')
    else:
        form = BatchCreationForm()
    return render(request, 'batchdb/create_batch.html', {'form': form})

@login_required
def update_batch(request, pk):
    batch = get_object_or_404(Batch, pk=pk)
    # Fetch active students in this batch via through model
    active_batch_students = BatchStudent.objects.filter(batch=batch, is_active=True).select_related('student')
    if request.method == 'POST':
        post_data = request.POST.copy()
        # Ensure trainers submit end_date (readonly field still must be present)
        if request.user.role == 'trainer' and not post_data.get('end_date'):
            # Use the existing batch end_date if no value is posted
            if batch.end_date:
                post_data['end_date'] = batch.end_date.isoformat()
        form = BatchUpdateForm(post_data, instance=batch)
        # Make end_date readonly for trainers at render time
        if request.user.role == 'trainer':
            form.fields['end_date'].widget.attrs['readonly'] = True
        if form.is_valid():
            batch = form.save(commit=False)
            # Enforce that trainers cannot change end_date server-side
            if request.user.role == 'trainer' and batch.end_date != Batch.objects.get(pk=batch.pk).end_date:
                batch.end_date = Batch.objects.get(pk=batch.pk).end_date

            batch.updated_by = request.user
            batch.save()
            form.save_m2m()

            # Process per-student updates: course_percentage and course_status
            updated_count = 0
            status_values = {choice[0] for choice in Student.COURSE_STATUS_CHOICES}
            for bs in active_batch_students:
                student = bs.student
                perc_key = f"student_{student.id}_percentage"
                status_key = f"student_{student.id}_status"
                perc_val = request.POST.get(perc_key)
                status_val = request.POST.get(status_key)

                changed = False
                # Update percentage if provided and valid
                if perc_val is not None and perc_val != "":
                    try:
                        perc_float = float(perc_val)
                        # Clamp between 0 and 100
                        if perc_float < 0:
                            perc_float = 0
                        if perc_float > 100:
                            perc_float = 100
                        if student.course_percentage != perc_float:
                            student.course_percentage = perc_float
                            changed = True
                    except ValueError:
                        # Ignore invalid percentage input
                        pass

                # Update status if provided and valid choice
                if status_val and status_val in status_values and student.course_status != status_val:
                    student.course_status = status_val
                    changed = True

                if changed:
                    student.save()
                    updated_count += 1

            if updated_count:
                messages.success(request, f"Updated {updated_count} student(s) for batch {batch.batch_id}.")
            messages.success(request, f"Batch {batch.batch_id} updated successfully.")
            return redirect('batchdb:batch_list')
    else:
        form = BatchUpdateForm(instance=batch, initial={'days': batch.days})
        if request.user.role == 'trainer':
            form.fields['end_date'].widget.attrs['readonly'] = True
    context = {
        'form': form,
        'batch': batch,
        'active_batch_students': active_batch_students,
        'course_status_choices': Student.COURSE_STATUS_CHOICES,
    }
    return render(request, 'batchdb/update_batch.html', context)

@login_required
def delete_batch(request, pk):
    batch = get_object_or_404(Batch, pk=pk)
    if request.method == 'POST':
        batch.delete()
        messages.success(request, "Batch deleted successfully.")
        return redirect('batchdb:batch_list')
    return render(request, 'batchdb/delete_confirm.html', {'batch': batch})

# AJAX Views
@login_required
def get_courses_by_category(request):
    category_id = request.GET.get('category_id')
    courses = Course.objects.filter(category_id=category_id).values('id', 'course_name')
    return JsonResponse(list(courses), safe=False)

@login_required
def get_trainers_for_course(request):
    course_id = request.GET.get('course_id')
    trainers = Trainer.objects.filter(stack__id=course_id).values('id','trainer_id','name')
    return JsonResponse(list(trainers), safe=False)

@login_required
def get_trainer_slots(request):
    trainer_id = request.GET.get('trainer_id')
    try:
        trainer = Trainer.objects.get(id=trainer_id)
        active_batches = Batch.objects.filter(trainer=trainer, batch_status__in=['IP', 'YTS'])
        taken_slots = [(batch.start_time, batch.end_time) for batch in active_batches]

        if not isinstance(trainer.timing_slots, list):
            return JsonResponse([], safe=False)

        available_slots = [
            slot for slot in trainer.timing_slots
            if isinstance(slot, dict) and
                (datetime.strptime(slot['start_time'], '%H:%M').time(), datetime.strptime(slot['end_time'], '%H:%M').time()) not in taken_slots
        ]
        
        formatted_slots = []
        for slot in available_slots:
            start_time = slot.get('start_time', '')
            end_time = slot.get('end_time', '')
            name = f"{datetime.strptime(start_time, '%H:%M').strftime('%I:%M %p')} - {datetime.strptime(end_time, '%H:%M').strftime('%I:%M %p')}"
            formatted_slots.append({'id': f"{start_time}-{end_time}", 'name': name})

        return JsonResponse(formatted_slots, safe=False)
    except Trainer.DoesNotExist:
        return JsonResponse([], safe=False)

@login_required
def get_students_for_course(request):
    course_id = request.GET.get('course_id')
    exclude_students_in_any_batch = request.GET.get('exclude_students_in_any_batch', 'false').lower() == 'true'

    # Fetch all students for this course with status YTS or IP
    students = Student.objects.filter(
        course_id=course_id,
        course_status__in=['YTS', 'IP']
    )

    if exclude_students_in_any_batch:
        # Get IDs of students already in active batches
        active_student_ids = BatchStudent.objects.filter(
            is_active=True
        ).values_list('student_id', flat=True)

        # Exclude them from the queryset
        students = students.exclude(id__in=active_student_ids)

    student_data = [
        {'id': s.id, 'student_id':s.student_id ,'first_name': s.first_name, 'last_name': s.last_name or ''}
        for s in students
    ]
    return JsonResponse(student_data, safe=False)

# API ViewSets

class BatchViewSet(viewsets.ModelViewSet):
    queryset = Batch.objects.all()
    serializer_class = BatchSerializer
    permission_classes = [IsBatchCoordinator | IsStaff | IsTrainer]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['batch_id', 'course__name', 'trainer__name']
    ordering_fields = ['start_date', 'end_date', 'created_at', 'batch_id']
    ordering = ['-created_at']
    
    def get_serializer_class(self):
        if self.action == 'retrieve':
            return BatchDetailSerializer
        return BatchSerializer
    
    def get_queryset(self):
        queryset = Batch.objects.all()
        
        # Filter by course category
        category_id = self.request.query_params.get('category_id', None)
        if category_id:
            queryset = queryset.filter(course__category_id=category_id)
        
        # Filter by course
        course_id = self.request.query_params.get('course_id', None)
        if course_id:
            queryset = queryset.filter(course_id=course_id)
        
        # Filter by trainer
        trainer_id = self.request.query_params.get('trainer_id', None)
        if trainer_id:
            queryset = queryset.filter(trainer_id=trainer_id)
        
        # Filter by batch status
        batch_status = self.request.query_params.get('batch_status', None)
        if batch_status:
            queryset = queryset.filter(batch_status=batch_status)
        
        # Filter by batch type
        batch_type = self.request.query_params.get('batch_type', None)
        if batch_type:
            queryset = queryset.filter(batch_type=batch_type)
        
        # Filter by date range
        start_date = self.request.query_params.get('start_date', None)
        end_date = self.request.query_params.get('end_date', None)
        
        if start_date:
            queryset = queryset.filter(start_date__gte=start_date)
        
        if end_date:
            queryset = queryset.filter(end_date__lte=end_date)
        
        return queryset
    
    @action(detail=True, methods=['get'])
    def students(self, request, pk=None):
        batch = self.get_object()
        batch_students = BatchStudent.objects.filter(batch=batch)
        
        # Filter by active status
        is_active = request.query_params.get('is_active', None)
        if is_active is not None:
            is_active = is_active.lower() == 'true'
            batch_students = batch_students.filter(is_active=is_active)
        
        serializer = BatchStudentSerializer(batch_students, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def add_student(self, request, pk=None):
       batch = self.get_object()
       student_ids = request.data.get('student_ids', [])

       if not student_ids:
           return Response({'error': 'Student IDs are required'}, status=status.HTTP_400_BAD_REQUEST)

       students_added = []
       errors = []

       for student_id in student_ids:
           try:
               student = Student.objects.get(pk=student_id)
           except Student.DoesNotExist:
               errors.append({'student_id': student_id, 'error': 'Student not found'})
               continue

           if BatchStudent.objects.filter(batch=batch, student=student, is_active=True).exists():
               errors.append({'student_id': student_id, 'error': 'Student is already active in this batch'})
               continue

           # Add student to batch
           batch_student, created = BatchStudent.objects.get_or_create(
               batch=batch,
               student=student,
               defaults={'is_active': True, 'activated_at': timezone.now()}
           )

           if not created and not batch_student.is_active:
               batch_student.activate(user=request.user)

           students_added.append(student)

       if errors:
           return Response({'errors': errors}, status=status.HTTP_400_BAD_REQUEST)

       return Response({'message': f'{len(students_added)} students added successfully'}, status=status.HTTP_200_OK)
            
    @action(detail=False, methods=['get'])
    def transactions(self, request):
        """Get transaction history for a batch"""
        batch_id = request.query_params.get('batch_id')
        
        if not batch_id:
            return Response({'error': 'Batch ID is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            batch = Batch.objects.get(pk=batch_id)
            transactions = BatchTransaction.objects.filter(batch=batch).order_by('-timestamp')
            
            # Filter by transaction type
            transaction_type = request.query_params.get('transaction_type', None)
            if transaction_type:
                transactions = transactions.filter(transaction_type=transaction_type)
            
            # Filter by date range
            start_date = request.query_params.get('start_date', None)
            if start_date:
                transactions = transactions.filter(timestamp__date__gte=start_date)
            
            end_date = request.query_params.get('end_date', None)
            if end_date:
                transactions = transactions.filter(timestamp__date__lte=end_date)
            
            page = self.paginate_queryset(transactions)
            if page is not None:
                serializer = BatchTransactionDetailSerializer(page, many=True)
                response = self.get_paginated_response(serializer.data)
                response.data['batch_id'] = batch.batch_id
                return response
            
            serializer = BatchTransactionDetailSerializer(transactions, many=True)
            return Response({
                'batch_id': batch.batch_id,
                'results': serializer.data
            })
        except Batch.DoesNotExist:
            return Response({'error': 'Batch not found'}, status=status.HTTP_404_NOT_FOUND)
    
    @action(detail=True, methods=['post'])
    def remove_student(self, request, pk=None):
        batch = self.get_object()
        student_id = request.data.get('student_id')
        
        if not student_id:
            return Response({'error': 'Student ID is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            batch_student = BatchStudent.objects.get(batch=batch, student_id=student_id, is_active=True)
        except BatchStudent.DoesNotExist:
            return Response({'error': 'Active student not found in this batch'}, status=status.HTTP_404_NOT_FOUND)
        
        # Deactivate the student
        batch_student.deactivate(user=request.user)
        
        return Response({'message': 'Student removed from batch'}, status=status.HTTP_200_OK)

class TransferRequestViewSet(viewsets.ModelViewSet):
    queryset = TransferRequest.objects.all()
    serializer_class = TransferRequestSerializer
    permission_classes = [IsBatchCoordinator | IsStaff]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['from_batch__batch_id', 'to_batch__batch_id']
    ordering_fields = ['requested_at', 'status']
    ordering = ['-requested_at']
    
    def perform_create(self, serializer):
        serializer.save(requested_by=self.request.user)

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return TransferRequestDetailSerializer
        return TransferRequestSerializer
    
    def get_queryset(self):
        queryset = TransferRequest.objects.all()
        
        # Filter by status
        status = self.request.query_params.get('status', None)
        if status:
            queryset = queryset.filter(status=status)
        
        # Filter by from_batch
        from_batch_id = self.request.query_params.get('from_batch', None)
        if from_batch_id:
            queryset = queryset.filter(from_batch_id=from_batch_id)
        
        # Filter by to_batch
        to_batch_id = self.request.query_params.get('to_batch', None)
        if to_batch_id:
            queryset = queryset.filter(to_batch_id=to_batch_id)
        
        # Filter by student
        student_id = self.request.query_params.get('student_id', None)
        if student_id:
            queryset = queryset.filter(students__id=student_id)
        
        return queryset
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        transfer_request = self.get_object()
        
        serializer = TransferRequestApprovalSerializer(
            data=request.data,
            context={'transfer_request': transfer_request}
        )
        
        if serializer.is_valid():
            approved_students = serializer.validated_data.get('approved_students', None)
            remarks = serializer.validated_data.get('remarks', None)
            
            with transaction.atomic():
                students = transfer_request.approve(
                    approved_by=request.user,
                    approved_students=approved_students,
                    remarks=remarks
                )
            
            return Response({
                'message': 'Transfer request approved successfully',
                'transferred_students': [str(student) for student in students]
            }, status=status.HTTP_200_OK)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        transfer_request = self.get_object()
        
        serializer = TransferRequestRejectionSerializer(
            data=request.data,
            context={'transfer_request': transfer_request}
        )
        
        if serializer.is_valid():
            remarks = serializer.validated_data.get('remarks', None)
            
            transfer_request.reject(
                rejected_by=request.user,
                remarks=remarks
            )
            
            return Response({
                'message': 'Transfer request rejected successfully'
            }, status=status.HTTP_200_OK)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class TrainerHandoverViewSet(viewsets.ModelViewSet):
    queryset = TrainerHandover.objects.all()
    serializer_class = TrainerHandoverSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['batch__batch_id', 'from_trainer__name', 'to_trainer__name']
    ordering_fields = ['requested_at', 'status']
    ordering = ['-requested_at']
    
    def perform_create(self, serializer):
        serializer.save(requested_by=self.request.user)

    def get_queryset(self):
        queryset = TrainerHandover.objects.all()
        
        # Filter by status
        status = self.request.query_params.get('status', None)
        if status:
            queryset = queryset.filter(status=status)
        
        # Filter by batch
        batch_id = self.request.query_params.get('batch_id', None)
        if batch_id:
            queryset = queryset.filter(batch_id=batch_id)
        
        # Filter by from_trainer
        from_trainer_id = self.request.query_params.get('from_trainer_id', None)
        if from_trainer_id:
            queryset = queryset.filter(from_trainer_id=from_trainer_id)
        
        # Filter by to_trainer
        to_trainer_id = self.request.query_params.get('to_trainer_id', None)
        if to_trainer_id:
            queryset = queryset.filter(to_trainer_id=to_trainer_id)
        
        return queryset
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        handover = self.get_object()
        
        serializer = TrainerHandoverApprovalSerializer(
            data=request.data,
            context={'handover': handover}
        )
        
        if serializer.is_valid():
            remarks = serializer.validated_data.get('remarks', None)
            
            with transaction.atomic():
                batch = handover.approve(
                    approved_by=request.user,
                    remarks=remarks
                )
            
            return Response({
                'message': 'Trainer handover approved successfully',
                'batch': batch.batch_id,
                'new_trainer': str(batch.trainer)
            }, status=status.HTTP_200_OK)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        handover = self.get_object()
        
        serializer = TrainerHandoverRejectionSerializer(
            data=request.data,
            context={'handover': handover}
        )
        
        if serializer.is_valid():
            remarks = serializer.validated_data.get('remarks', None)
            
            handover.reject(
                rejected_by=request.user,
                remarks=remarks
            )
            
            return Response({
                'message': 'Trainer handover rejected successfully'
            }, status=status.HTTP_200_OK)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class BatchTransactionViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = BatchTransaction.objects.all()
    serializer_class = BatchTransactionSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['batch__batch_id', 'transaction_type']
    ordering_fields = ['timestamp']
    ordering = ['-timestamp']
    
    def get_serializer_class(self):
        return BatchTransactionDetailSerializer
    
    def get_queryset(self):
        queryset = BatchTransaction.objects.all()
        
        # Filter by batch
        batch_id = self.request.query_params.get('batch_id', None)
        if batch_id:
            queryset = queryset.filter(batch_id=batch_id)
        
        # Filter by transaction type
        transaction_type = self.request.query_params.get('transaction_type', None)
        if transaction_type:
            queryset = queryset.filter(transaction_type=transaction_type)
        
        # Filter by student
        student_id = self.request.query_params.get('student_id', None)
        if student_id:
            queryset = queryset.filter(affected_students__id=student_id)
        
        # Filter by date range
        start_date = self.request.query_params.get('start_date', None)
        end_date = self.request.query_params.get('end_date', None)
        
        if start_date:
            queryset = queryset.filter(timestamp__date__gte=start_date)
        
        if end_date:
            queryset = queryset.filter(timestamp__date__lte=end_date)
        
        return queryset

class StudentHistoryViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    def list(self, request):
        student_id = request.query_params.get('student_id')
        if not student_id:
            return Response({'error': 'Student ID is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            student = Student.objects.get(pk=student_id)
        except Student.DoesNotExist:
            return Response({'error': 'Student not found'}, status=status.HTTP_404_NOT_FOUND)

        history_data = student.get_batch_history()
        return Response(history_data)

# Student History Report
def student_batch_history(request):
    """View to display a student's batch history"""
    student_id = request.GET.get('student_id')
    
    if not student_id:
        messages.error(request, "Student ID is required.")
        return redirect('batchdb:batch_list')

    try:
        student = Student.objects.get(pk=student_id)
    except Student.DoesNotExist:
        messages.error(request, "Student not found.")
        return redirect('batchdb:batch_list')

    history_data = BatchStudent.get_student_batch_history(student)
    
    transactions = history_data['transactions']
    paginator = Paginator(transactions, 10)
    page = request.GET.get('page')
    
    try:
        transactions = paginator.page(page)
    except PageNotAnInteger:
        transactions = paginator.page(1)
    except EmptyPage:
        transactions = paginator.page(paginator.num_pages)
        
    for transaction in transactions:
        details = transaction.details
        if isinstance(details, str):
            try:
                details = json.loads(details)
            except json.JSONDecodeError:
                details = {}
        transaction.details = details
        
    history_data['transactions'] = transactions
    
    context = {
        'student': student,
        'student_data': history_data,
    }
    
    return render(request, 'batchdb/student_history.html', context)

@require_GET
def get_students_for_batch(request):
    """AJAX view to get students for a batch"""
    batch_id = request.GET.get('batch_id')
    
    if not batch_id:
        return JsonResponse({'error': 'Batch ID is required'}, status=400)
    
    try:
        batch = Batch.objects.get(pk=batch_id)
        students = batch.batchstudent_set.all()
        
        student_data = [{
            'id': student.student.id,
            'student_id': student.student.student_id,
            'name': f"{student.student.first_name} {student.student.last_name or ''}".strip(),
            'email': student.student.email,
            'phone': student.student.phone,
            'is_active': student.is_active
        } for student in students]
        
        return JsonResponse({
            'batch_id': batch.batch_id,
            'students': student_data
        })
    except Batch.DoesNotExist:
        return JsonResponse({'error': 'Batch not found'}, status=404)

# AJAX endpoints for cascading selects
@require_GET
def get_courses_by_category(request):
    category_id = request.GET.get('category_id')
    
    if not category_id:
        return JsonResponse({'error': 'Category ID is required'}, status=400)
    
    courses = Course.objects.filter(category_id=category_id).values('id', 'course_name', 'code')
    return JsonResponse(list(courses), safe=False)

@require_GET
def get_trainers_by_course(request):
    course_id = request.GET.get('course_id')
    
    if not course_id:
        return JsonResponse({'error': 'Course ID is required'}, status=400)
    
    # Get trainers who can teach this course
    trainers = Trainer.objects.filter(courses=course_id).values('id', 'name', 'email', 'phone')
    return JsonResponse(list(trainers), safe=False)

@require_GET
def get_students_by_course(request):
    course_id = request.GET.get('course_id')
    batch_id = request.GET.get('exclude_batch_id')  # Optional: exclude students already in a specific batch
    
    if not course_id:
        return JsonResponse({'error': 'Course ID is required'}, status=400)
    
    # Get students who are enrolled in this course
    students_query = Student.objects.filter(enrolled_course_id=course_id)
    
    # Exclude students who are already active in the specified batch
    if batch_id:
        active_student_ids = BatchStudent.objects.filter(
            batch_id=batch_id,
            is_active=True
        ).values_list('student_id', flat=True)
        
        students_query = students_query.exclude(id__in=active_student_ids)
    
    students = students_query.values('id', 'name', 'email', 'phone')
    return JsonResponse(list(students), safe=False)

@login_required
def get_students_not_in_batch(request):
   """
   AJAX view to get students who are not in any active batch.
   """
   search_term = request.GET.get('q', '')
   
   # Get IDs of students who are already in an active batch
   active_student_ids = BatchStudent.objects.filter(is_active=True).values_list('student_id', flat=True)
   
   # Filter students who are not in the active list
   students = Student.objects.exclude(id__in=active_student_ids)
   
   if search_term:
       students = students.filter(
           Q(first_name__icontains=search_term) | Q(last_name__icontains=search_term) | Q(student_id__icontains=search_term)
       )
       
   student_data = list(students.values('id','student_id', 'first_name', 'last_name'))
   return JsonResponse(student_data, safe=False)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def available_students_for_transfer(request):
    batch_id = request.query_params.get('batch_id')
    if not batch_id:
        return Response({'error': 'Batch ID is required'}, status=status.HTTP_400_BAD_REQUEST)
    
    students = Student.objects.filter(batchstudent__batch_id=batch_id, batchstudent__is_active=True)
    serializer = StudentSerializer(students, many=True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def available_batches_for_transfer(request):
    from_batch_id = request.query_params.get('from_batch_id')
    if not from_batch_id:
        return Response({'error': 'From Batch ID is required'}, status=status.HTTP_400_BAD_REQUEST)

    # Fetch all batches and exclude the source batch
    available_batches = Batch.objects.exclude(id=from_batch_id)
    serializer = BatchSerializer(available_batches, many=True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def available_trainers_for_handover(request):
    batch_id = request.query_params.get('batch_id')
    if not batch_id:
        return Response({'error': 'Batch ID is required'}, status=status.HTTP_400_BAD_REQUEST)
    
    batch = get_object_or_404(Batch, pk=batch_id)
    
    # Exclude the current trainer from the list of available trainers
    available_trainers = Trainer.objects.exclude(id=batch.trainer.id)
    serializer = TrainerSerializer(available_trainers, many=True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def available_batches_for_handover(request):
    trainer_id = request.query_params.get('trainer_id')
    if not trainer_id:
        return Response({'error': 'Trainer ID is required'}, status=status.HTTP_400_BAD_REQUEST)
    
    # Get all batches assigned to the specified trainer
    batches = Batch.objects.filter(trainer_id=trainer_id)
    serializer = BatchSerializer(batches, many=True)
    return Response(serializer.data)

@login_required
def download_batch_template(request):
    data = {
        'batch_id': ['BAT_101', ''],
        'module_name': ['Python', 'Java'],
        'batch_type': ['Weekday', 'Weekend'],
        'trainer': ['Trainer A', 'Trainer B'],
        'start_date': ['2025-08-01', '2025-08-01'],
        'end_date': ['2025-10-01', '2025-10-01'],
        'time_slot': ['9:00 AM - 10:30 AM', '7:00 PM - 8:30 PM'],
        'students': ['BTR0001,BTR0002', 'BTR0003,BTR0004']
    }
    df = pd.DataFrame(data)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="batch_template.xlsx"'
    df.to_excel(response, index=False)

    return response

@login_required
def import_batches(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "No batch file was uploaded.")
            return redirect('batchdb:batch_list')

        try:
            df = pd.read_excel(excel_file)
        except Exception as e:
            messages.error(request, f"Error reading batch Excel file: {e}")
            return redirect('batchdb:batch_list')

        required_columns = [
            'module_name', 'batch_type', 'trainer', 'start_date', 'end_date', 'time_slot', 'students'
        ]
        if not all(col in df.columns for col in required_columns):
            messages.error(request, f"Batch Excel file must contain the following columns: {', '.join(required_columns)}")
            return redirect('batchdb:batch_list')

        error_rows = []
        success_count = 0

        with transaction.atomic():
            for index, row in df.iterrows():
                try:
                    module_name = row['module_name']
                    batch_type = row['batch_type']
                    trainer_name = row['trainer']
                    start_date_str = str(row['start_date']).split(' ')[0]
                    end_date_str = str(row['end_date']).split(' ')[0]
                    time_slot = row['time_slot']
                    student_ids_str = row['students']

                    if not all([module_name, batch_type, trainer_name, start_date_str, end_date_str, time_slot, student_ids_str]):
                        raise ValueError("All fields are mandatory.")
                    
                    student_ids = str(student_ids_str).split(',')

                    try:
                        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                    except ValueError:
                        raise ValueError("Invalid date format. Use YYYY-MM-DD.")

                    trainer, _ = Trainer.objects.get_or_create(name=trainer_name)
                    
                    batch_id_from_sheet = row.get('batch_id')
                    if pd.notna(batch_id_from_sheet) and str(batch_id_from_sheet).strip():
                        batch_id = str(batch_id_from_sheet).strip()
                        if Batch.objects.filter(batch_id=batch_id).exists():
                            raise ValueError(f"Batch with ID '{batch_id}' already exists.")
                        batch = Batch(batch_id=batch_id)
                    else:
                        batch = Batch()

                    batch.module_name = module_name
                    batch.batch_type = batch_type
                    batch.trainer = trainer
                    batch.start_date = start_date
                    batch.end_date = end_date
                    batch.time_slot = time_slot
                    batch.save()

                    students = []
                    for student_id in student_ids:
                        student_id = student_id.strip()
                        try:
                            student = Student.objects.get(student_id=student_id)
                            students.append(student)
                        except Student.DoesNotExist:
                            raise ValueError(f"Student with ID {student_id} not found.")
                    
                    batch.students.set(students)

                    for student in students:
                        if student.pl_required:
                            Placement.objects.get_or_create(student=student)

                    success_count += 1

                except Exception as e:
                    error_row = row.to_dict()
                    error_row['error_reason'] = str(e)
                    error_rows.append(error_row)

        if error_rows:
            request.session['error_rows_batch'] = error_rows
            messages.warning(request, f"Successfully created {success_count} batches. {len(error_rows)} records had errors.")
            return redirect('download_error_report_batch')
        else:
            messages.success(request, f"Successfully created {success_count} batches.")
            return redirect('batchdb:batch_list')

    return render(request, 'batchdb/import_batches.html')

@login_required
def download_error_report_batch(request):
    error_rows = request.session.get('error_rows_batch', [])
    if not error_rows:
        messages.error(request, "No error report to download.")
        return redirect('batchdb:batch_list')

    df = pd.DataFrame(error_rows)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="batch_error_report.csv"'
    df.to_csv(response, index=False)

    del request.session['error_rows_batch']

    return response

@login_required
def batch_report(request, pk):
    batch = get_object_or_404(Batch, pk=pk)
    
    # Get all students who have ever been in the batch
    all_batch_students = BatchStudent.objects.filter(batch=batch).select_related('student').order_by('student__first_name')
    
    # Separate active and inactive students
    active_students = [bs.student for bs in all_batch_students if bs.is_active]
    inactive_students = []
    for bs in all_batch_students:
        if not bs.is_active:
            student = bs.student
            # Find the transfer reason
            transfer_request = TransferRequest.objects.filter(
                from_batch=batch,
                students=student
            ).order_by('-requested_at').first()
            student.transfer_reason = transfer_request.remarks if transfer_request else "N/A"
            inactive_students.append(student)
    
    # Get all transactions for the batch
    transactions = BatchTransaction.objects.filter(batch=batch).select_related('user').order_by('-timestamp')
    
    # Get trainer handover history
    handover_history = TrainerHandover.objects.filter(batch=batch, status='APPROVED').order_by('-approved_at')

    # Paginate transactions
    paginator = Paginator(transactions, 10)
    page = request.GET.get('page')
    
    try:
        transactions_page = paginator.page(page)
    except PageNotAnInteger:
        transactions_page = paginator.page(1)
    except EmptyPage:
        transactions_page = paginator.page(paginator.num_pages)
        
    context = {
        'batch': batch,
        'active_students': active_students,
        'inactive_students': inactive_students,
        'transactions': transactions_page,
        'handover_history': handover_history,
    }
    
    return render(request, 'batchdb/batch_report.html', context)

def view_handover_requests(request):
    handover_requests = TrainerHandover.objects.filter(status='PENDING')
    return render(request, 'batchdb/handover_requests.html', {'requests': handover_requests})

def update_handover_status(request, pk):
    handover = get_object_or_404(TrainerHandover, pk=pk)
    if request.method == 'POST':
        status = request.POST.get('status')
        if status == 'approve':
            handover.approve(request.user)
        elif status == 'reject':
            handover.reject(request.user)
    return redirect('batchdb:view_handover_requests')

class RequestListView(LoginRequiredMixin, ListView):
    """
    View for displaying and managing all transfer and handover requests in one place
    """
    template_name = 'batchdb/requests_list.html'
    context_object_name = 'requests'
    paginate_by = 10

    def get_queryset(self):
        # Automatically expire requests before filtering
        TransferRequest.expire_pending_requests()
        TrainerHandover.expire_pending_requests()

        # Get filter parameters
        request_type = self.request.GET.get('request_type', '')
        status = self.request.GET.get('status', '')
        from_date = self.request.GET.get('from_date', '')
        to_date = self.request.GET.get('to_date', '')
        search_query = self.request.GET.get('q', '')

        # Initialize empty querysets
        transfer_requests = TransferRequest.objects.select_related(
            'from_batch', 'to_batch', 'requested_by'
        ).prefetch_related('students').all()
        handover_requests = TrainerHandover.objects.select_related(
            'batch', 'from_trainer', 'to_trainer', 'requested_by'
        ).all()

        # Apply filters to both querysets
        if status:
            transfer_requests = transfer_requests.filter(status=status)
            handover_requests = handover_requests.filter(status=status)
        if from_date:
            try:
                from_date = datetime.strptime(from_date, '%Y-%m-%d')
                transfer_requests = transfer_requests.filter(requested_at__gte=from_date)
                handover_requests = handover_requests.filter(requested_at__gte=from_date)
            except (ValueError, TypeError):
                pass # Ignore invalid date format

        if to_date:
            try:
                to_date = datetime.strptime(to_date, '%Y-%m-%d')
                to_date = to_date.replace(hour=23, minute=59, second=59)
                transfer_requests = transfer_requests.filter(requested_at__lte=to_date)
                handover_requests = handover_requests.filter(requested_at__lte=to_date)
            except (ValueError, TypeError):
                pass # Ignore invalid date format

        if search_query:
            transfer_requests = transfer_requests.filter(
                Q(from_batch__batch_id__icontains=search_query) |
                Q(to_batch__batch_id__icontains=search_query) |
                Q(students__first_name__icontains=search_query) |
                Q(students__last_name__icontains=search_query) |
                Q(requested_by__name__icontains=search_query)
            ).distinct()
            
            handover_requests = handover_requests.filter(
                Q(batch__batch_id__icontains=search_query) |
                Q(from_trainer__name__icontains=search_query) |
                Q(to_trainer__name__icontains=search_query) |
                Q(requested_by__name__icontains=search_query)
            ).distinct()

        # Prepare and combine requests
        combined_requests = []
        if request_type != 'handover':
            for req in transfer_requests:
                req.request_type = 'transfer'
                combined_requests.append(req)
        
        if request_type != 'transfer':
            for req in handover_requests:
                req.request_type = 'handover'
                combined_requests.append(req)

        # Sort by requested_at (newest first)
        combined_requests.sort(key=lambda x: x.requested_at, reverse=True)
        
        return combined_requests

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['requests'] = context['page_obj']
        return context
        
@login_required
def request_details(request, request_id):
    """API endpoint to get details of a request"""
    try:
        transfer_request = TransferRequest.objects.select_related(
            'from_batch', 'to_batch', 'requested_by', 'approved_by'
        ).prefetch_related('students').get(id=request_id)
        
        students_info = ", ".join([f"{s.first_name} {s.last_name or ''}".strip() for s in transfer_request.students.all()])
        data = {
            'id': transfer_request.id,
            'request_type': 'Student Transfer',
            'status': transfer_request.get_status_display(),
            'students': students_info,
            'students_count': transfer_request.students.count(),
            'from_batch': transfer_request.from_batch.batch_id,
            'to_batch': transfer_request.to_batch.batch_id,
            'requested_by': transfer_request.requested_by.name,
            'requested_at': transfer_request.requested_at.strftime('%d %b %Y'),
            'remarks': transfer_request.remarks
        }
        if transfer_request.status == 'APPROVED':
            data['approved_by'] = transfer_request.approved_by.name if transfer_request.approved_by else 'System'
            data['approved_at'] = transfer_request.approved_at.strftime('%d %b %Y') if transfer_request.approved_at else ''
        elif transfer_request.status == 'REJECTED':
            data['rejected_by'] = transfer_request.approved_by.name if transfer_request.approved_by else 'System'
            data['rejected_at'] = transfer_request.approved_at.strftime('%d %b %Y') if transfer_request.approved_at else ''
        return JsonResponse(data)
    except TransferRequest.DoesNotExist:
        pass

    try:
        handover_request = TrainerHandover.objects.select_related(
            'batch', 'from_trainer', 'to_trainer', 'requested_by', 'approved_by'
        ).get(id=request_id)
        data = {
            'id': handover_request.id,
            'request_type': 'Trainer Handover',
            'status': handover_request.get_status_display(),
            'batch': handover_request.batch.batch_id,
            'from_trainer': handover_request.from_trainer.name,
            'to_trainer': handover_request.to_trainer.name,
            'requested_by': handover_request.requested_by.name,
            'requested_at': handover_request.requested_at.strftime('%d %b %Y'),
            'remarks': handover_request.remarks
        }
        if handover_request.status == 'APPROVED':
            data['approved_by'] = handover_request.approved_by.name if handover_request.approved_by else 'System'
            data['approved_at'] = handover_request.approved_at.strftime('%d %b %Y') if handover_request.approved_at else ''
        elif handover_request.status == 'REJECTED':
            data['rejected_by'] = handover_request.approved_by.name if handover_request.approved_by else 'System'
            data['rejected_at'] = handover_request.approved_at.strftime('%d %b %Y') if handover_request.approved_at else ''
        return JsonResponse(data)
    except TrainerHandover.DoesNotExist:
        pass

    return JsonResponse({'error': 'Request not found'}, status=404)

@login_required
@require_POST
def approve_request(request, request_id):
    """API endpoint to approve a request"""
    try:
        data = json.loads(request.body)
        remarks = data.get('remarks')
        approved_student_ids = data.get('approved_students')
    except json.JSONDecodeError:
        remarks = None
        approved_student_ids = None

    approved_students = None
    if approved_student_ids:
        try:
            approved_students = Student.objects.filter(id__in=approved_student_ids)
        except (ValueError, TypeError):
            return JsonResponse({'error': 'Invalid student IDs provided.'}, status=400)

    try:
        transfer_request = TransferRequest.objects.get(id=request_id, status='PENDING')
        transfer_request.approve(
            approved_by=request.user,
            remarks=remarks,
            approved_students=approved_students
        )
        messages.success(request, "Transfer request approved successfully.")
        return JsonResponse({'success': True, 'message': 'Transfer request approved successfully.'})
    except TransferRequest.DoesNotExist:
        pass
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

    try:
        handover_request = TrainerHandover.objects.get(id=request_id, status='PENDING')
        handover_request.approve(approved_by=request.user, remarks=remarks)
        messages.success(request, "Handover request approved successfully.")
        return JsonResponse({'success': True, 'message': 'Handover request approved successfully.'})
    except TrainerHandover.DoesNotExist:
        pass
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'error': 'Pending request not found'}, status=404)

@login_required
@require_POST
def reject_request(request, request_id):
    """API endpoint to reject a request"""
    try:
        data = json.loads(request.body)
        reason = data.get('remarks', 'Rejected without a specific reason.')
    except json.JSONDecodeError:
        reason = 'Rejected without a specific reason.'

    try:
        transfer_request = TransferRequest.objects.get(id=request_id, status='PENDING')
        transfer_request.reject(rejected_by=request.user, remarks=reason)
        messages.success(request, "Transfer request rejected successfully.")
        return JsonResponse({'success': True, 'message': 'Transfer request rejected successfully.'})
    except TransferRequest.DoesNotExist:
        pass
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

    try:
        handover_request = TrainerHandover.objects.get(id=request_id, status='PENDING')
        handover_request.reject(rejected_by=request.user, remarks=reason)
        messages.success(request, "Handover request rejected successfully.")
        return JsonResponse({'success': True, 'message': 'Handover request rejected successfully.'})
    except TrainerHandover.DoesNotExist:
        pass
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'error': 'Pending request not found'}, status=404)

@login_required
def export_batch_data(request, batch_id):
    try:
        batch = Batch.objects.get(id=batch_id)
    except Batch.DoesNotExist:
        messages.error(request, "Batch not found.")
        return redirect('batchdb:batch_list')

    # Prepare data for export
    data = {
        'Batch ID': [batch.batch_id],
        'Course': [batch.course.course_name if batch.course else 'N/A'],
        'Trainer': [batch.trainer.name if batch.trainer else 'N/A'],
        'Start Date': [batch.start_date.strftime('%d-%m-%Y')],
        'End Date': [batch.end_date.strftime('%d-%m-%Y')],
        'Slot Time': [batch.get_slottime],
        'Status': [batch.get_batch_status_display()],
        'Days': [', '.join(batch.days)],
    }

    # Student details
    students = batch.students.all()
    if students:
        student_data = {
            'Student ID': [s.student_id for s in students],
            'Student Name': [f"{s.first_name} {s.last_name or ''}" for s in students],
            'Email': [s.email for s in students],
            'Phone': [s.phone for s in students],
        }
        max_len = max(len(v) for v in student_data.values())
        for k, v in data.items():
            v.extend([''] * (max_len - len(v)))
        data.update(student_data)

    # Create a DataFrame
    df = pd.DataFrame(data)

    # Create an Excel response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{batch.batch_id}_details.xlsx"'
    df.to_excel(response, index=False)

    return response

@login_required
def export_all_batches_data(request):
    batches = Batch.objects.all().order_by('batch_id')
    
    if not batches.exists():
        messages.error(request, "No batches to export.")
        return redirect('batchdb:batch_list')

    all_batches_data = []
    batch_row_map = {}
    current_row = 1

    for batch in batches:
        batch_info = {
            'Batch ID': batch.batch_id,
            'Course': batch.course.course_name if batch.course else 'N/A',
            'Trainer': batch.trainer.name if batch.trainer else 'N/A',
            'Start Date': batch.start_date.strftime('%d-%m-%Y'),
            'End Date': batch.end_date.strftime('%d-%m-%Y'),
            'Slot Time': batch.get_slottime,
            'Status': batch.get_batch_status_display(),
            'Days': ', '.join(batch.days),
        }

        students = batch.students.all()
        start_row = current_row
        
        if students:
            for student in students:
                student_info = batch_info.copy()
                student_info.update({
                    'Student ID': student.student_id,
                    'Student Name': f"{student.first_name} {student.last_name or ''}",
                    'Email': student.email,
                    'Phone': student.phone,
                })
                all_batches_data.append(student_info)
                current_row += 1
        else:
            all_batches_data.append(batch_info)
            current_row += 1
            
        batch_row_map[batch.id] = (start_row, current_row - 1)

    df = pd.DataFrame(all_batches_data)
    column_order = ['Batch ID', 'Course', 'Trainer', 'Start Date', 'End Date', 'Slot Time', 'Status', 'Days', 'Student ID', 'Student Name', 'Email', 'Phone']
    df = df.reindex(columns=column_order)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='All Batches')
        workbook = writer.book
        worksheet = writer.sheets['All Batches']

        colors = ['FFFFCC', 'CCFFCC', 'CCFFFF', 'FFCCFF', 'CCE5FF', 'FFDDAA']
        color_index = 0

        for batch_id, (start_row, end_row) in batch_row_map.items():
            fill_color = PatternFill(start_color=colors[color_index % len(colors)],
                                     end_color=colors[color_index % len(colors)],
                                     fill_type="solid")
            for row in range(start_row + 1, end_row + 2):
                for col in range(1, len(column_order) + 1):
                    worksheet.cell(row=row, column=col).fill = fill_color
            color_index += 1

    output.seek(0)
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="all_batches_details.xlsx"'
    return response
